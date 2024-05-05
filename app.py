import openpyxl

## We need to start with a workbook object, we can even create an empty
## workbook in memory, or load and existing work book on disc. So, this
## module openpyxl has this workbook class and you can create a new workbook
## object.

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a1"]
print(cell.value)
# cell.value = 1
print(cell.row)
print(cell.column)
print(cell.coordinate)

print(sheet.cell(row=1, column=1))
print(sheet.max_row)
print(sheet.max_column)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row=row, column=column)
#         print(cell.value)

cells = sheet["a:c"]
print(cells)

print(sheet["a1:c3"])
sheet[1:3]

sheet.append([1, 2, 3])
# sheet.insert_rows([])
# sheet.insert_cols([])
# sheet.delete_rows([])
# sheet.delete_cols([])

wb.save("transactions2.xlsx")

print("##################################")

## Command Query Seperation: This principle states that our methods or
## functions should either be commands that perform an action to change
## the state of a system, or queries that return an answer to the caller
## without changing the state or causing side effects. So our methods
## should wither be commands or queries, but not both.

## wb.create_sheet() This is an example of a commanf method
## because it's responsible for performing a task. The task of creating
## a sheet. As a result of calling this method, the state of our system,
## (our workbook) changes, so everytime we call it, we get a new sheet in
## this woekbook..

## sheet.cell() This is an example of a query method. We use it to access
## a given cell. However, this method violates the command query
## seperation.
## Reason:
## We are going to print the value of all the sales in the first column
for row in range(1, 10):
    cell = sheet.cell(row, 1) ## This cell method will create more 6 rows
    ## because that we have just 4 rows and we loop on 10
    print(cell.value)
sheet.append([1, 4, 5])
wb.save("transaction3.xlsx")
## This is the violation of command query seperation.
